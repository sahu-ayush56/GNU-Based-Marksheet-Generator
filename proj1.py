# Ayush Sahu - 1901CB13
# Aditya Goyal - 1901EE06
import genericpath
import re
from typing import final
from flask import Flask
from flask import render_template, url_for, request, redirect
import smtplib
import shutil
from werkzeug.wrappers import response
from os import name
import os.path
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
from openpyxl.drawing.image import Image  
from collections import defaultdict
import json
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
app = Flask(__name__)
personal_info = os.path.join("templates","config.json")
with open(personal_info, 'r') as c:
    params = json.load(c)["params"]

path1 = "sample_input"
dir1 = os.path.join(path1)
if os.path.isdir(dir1) == True:
    shutil.rmtree(dir1)
os.mkdir(path1)
path2 = "sample_output"
dir2 = os.path.join(path2)
if os.path.isdir(dir2) == True:
    shutil.rmtree(dir2)

loc = os.path.join(app.root_path,path1)   
# print(loc) 
app.config["input_file_loc"] = loc
ndir, master_flag , response_flag, pos, neg,ldir,btn1,btn2,btn3,notlast,cantsend,gen_pos,gen_neg = (0,)*13
flag1,btn1_flag,btn2_flag= (2,)*3
gen_list = []

direc="sample_output"

response_path = os.path.join("sample_input","responses.csv")
master_path=os.path.join("sample_input","master_roll.csv")

def generate_marksheet(right_per,wrong_per):
    d = defaultdict(list)
    ans=[]
    dirpath = os.path.join(direc)
    #If output folder is present then removing it and then making it again for avoiding readding of data
        
    if os.path.isdir(dirpath) == False:
        os.mkdir(direc)
    final_srt=os.path.join(direc,"marksheet")
    if os.path.isdir(final_srt)==False:
        os.mkdir(final_srt)  
    else:
        marksheet = os.listdir(final_srt)
        if "concise_marksheet.csv" not in marksheet:
            shutil.rmtree(final_srt)
            os.mkdir(final_srt)
        else:
            concise1 = os.path.join(final_srt,"concise_marksheet.csv")
            concise2 = os.path.join(dirpath,"concise_marksheet.csv")
            os.replace(concise1,concise2)
            shutil.rmtree(final_srt)
            os.mkdir(final_srt)
            os.replace(concise2,concise1)
    with open(master_path,'r') as f:
        reader=csv.reader(f)
        for words in reader:
            if words[0]!="roll":
              d[words[0]].append(False)
              d[words[0]].append(words[1])
    
    with open(response_path,'r') as f:
        reader =csv.reader(f)
        flag=0
        for words in reader:
            if words[6]=="ANSWER":
                flag=1
                ans=words[7:]
                break

    # if no row with roll_num="ANSWER" exist
    if flag==0:  
        print(flag)          
        raise ValueError

    with open(response_path,'r') as f:
        global ft3, black_border
        reader =csv.reader(f)
        for words in reader: 
            right=0
            wrong=0
            not_attempt=0
            max=0
            roll_no=words[6]
            roll_no.upper()
            if roll_no=="Roll Number":
                continue
            check=os.path.join(final_srt,roll_no+".xlsx")
            d[roll_no][0]=True
            d[roll_no].append(words[1])
            d[roll_no].append(words[4])
            if os.path.isfile(check) is False:
                wb = Workbook()
                sheet = wb.active
                sheet.title="quiz"
                sheet.column_dimensions["A"].width = 18
                sheet.column_dimensions["B"].width = 18
                sheet.column_dimensions["C"].width = 18
                sheet.column_dimensions["D"].width = 18
                sheet.column_dimensions["E"].width = 18
                img = Image('photo.png')
                img.height=80
                img.width=615
                sheet.add_image(img,"A1")
                sheet.merge_cells('A5:E5')  
                cell = sheet["A5"]  
                cell.value = 'Mark Sheet'  
                ft = Font(name='Century', size=18,bold=True,underline="single")
                ft1 = Font(name='Century', size=12,bold=True)
                ft2 = Font(name='Century', size=12)
                cell.font = ft
                sheet.append(["Name:",words[3],"","Exam:","quiz"])
                sheet.append(["Roll Numer:",roll_no])
                alignment=Alignment(horizontal="right")
                alignment1=Alignment(horizontal="center")
                sheet["B6"].font=ft1 
                sheet["B7"].font=ft1
                sheet["E6"].font=ft1
                sheet["A6"].alignment=alignment
                sheet["A7"].alignment=alignment
                sheet["D6"].alignment=alignment
                sheet["A6"].font=ft2 
                sheet["A7"].font=ft2
                sheet["D6"].font=ft2
                cell.alignment = Alignment(horizontal='center', vertical='center')  
                sheet.append([])
                sheet.append(["","Right","Wrong","Not Attempt","MAX"])
                black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for label in ["E", "B", "C", "D","A"]:
                    idx = label + str(9)
                    sheet[idx].alignment = alignment1
                    sheet[idx].font = ft1
                    sheet[idx].border=black_border
                sheet["A15"]="Student Ans"
                sheet["B15"]="Correct Ans"
                sheet["D15"]="Student Ans"
                sheet["E15"]="Correct Ans"
                for label in ["A", "B", "E", "D"]:
                    idx = label + str(15)
                    sheet[idx].alignment = alignment1
                    sheet[idx].font = ft1
                    sheet[idx].border=black_border
                for i in range(len(ans)):
                    max=max+1
                    if words[i+7]=="":
                        not_attempt=not_attempt+1
                    elif ans[i]==words[i+7]:
                        right=right+1
                        ft3 = Font(name='Century', size=12,color="00008000")
                    else:
                        wrong=wrong+1
                        ft3 = Font(name='Century', size=12,color="00FF0000")
                    t1=""
                    t2="" 
                    com=0
                    if i<25:
                        t1="B"
                        t2="A"
                        com=i
                    else:
                        t1="E"
                        t2="D"
                        com=i-25
                    sheet[t1+str(com+16)]=ans[i]
                    sheet[t2+str(com+16)]=words[i+7]
                    sheet[t1+str(com+16)].alignment = alignment1
                    sheet[t1+str(com+16)].border=black_border
                    sheet[t1+str(com+16)].font=Font(name='Century', size=12,color="000000FF")
                    sheet[t2+str(com+16)].alignment = alignment1
                    sheet[t2+str(com+16)].border=black_border
                    sheet[t2+str(com+16)].font=ft3
                for i in range(3):
                    lis=[]
                    if(i==0):
                        lis=["No.",right,wrong,not_attempt,max]
                    elif(i==1):
                        lis=["Marking",right_per,wrong_per,0,""]
                    elif(i==2):
                        lis=["Total",right_per*right,wrong_per*wrong,"",str((right_per*right+wrong_per*wrong))+"/"+str(max*right_per)]
                    for label in ["A", "B", "C", "D","E"]:
                        sheet[label+str(i+10)]=lis[ord(label)-65]
                        sheet[label+str(i+10)].alignment = alignment1
                        sheet[label+str(i+10)].border=black_border
                sheet["A10"].font=ft1
                sheet["A11"].font=ft1
                sheet["A12"].font=ft1
                sheet["D10"].font=ft1
                sheet["D11"].font=ft1
                sheet["E10"].font=ft1
                sheet["B10"].font=Font(name='Century', size=12,color="00008000")
                sheet["B11"].font=Font(name='Century', size=12,color="00008000")
                sheet["B12"].font=Font(name='Century', size=12,color="00008000")
                sheet["C10"].font=Font(name='Century', size=12,color="00FF0000")
                sheet["C11"].font=Font(name='Century', size=12,color="00FF0000")
                sheet["C12"].font=Font(name='Century', size=12,color="00FF0000")
                sheet["E12"].font=Font(name='Century', size=12,color="000000FF")
                wb.save(check)
    
    for key, value in d.items():
        if value[0]==False:
            check=os.path.join(final_srt,key+".xlsx")
            if os.path.isfile(check) is False:
                wb = Workbook()
                sheet = wb.active
                sheet.title="quiz"
                sheet.column_dimensions["A"].width = 18
                sheet.column_dimensions["B"].width = 18
                sheet.column_dimensions["C"].width = 18
                sheet.column_dimensions["D"].width = 18
                sheet.column_dimensions["E"].width = 18
                img = Image('photo.png')
                img.height=80
                img.width=615
                sheet.add_image(img,"A1")
                sheet.merge_cells('A5:E5')  
                cell = sheet["A5"]  
                cell.value = 'Mark Sheet'  
                ft = Font(name='Century', size=18,bold=True,underline="single")
                ft1 = Font(name='Century', size=12,bold=True)
                ft2 = Font(name='Century', size=12)
                cell.font = ft
                sheet.append(["Name:",value[1],"","Exam:","quiz"])
                sheet.append(["Roll Numer:",key])
                alignment=Alignment(horizontal="right")
                alignment1=Alignment(horizontal="center")
                sheet["B6"].font=ft1 
                sheet["B7"].font=ft1
                sheet["E6"].font=ft1
                sheet["A6"].alignment=alignment
                sheet["A7"].alignment=alignment
                sheet["D6"].alignment=alignment
                sheet["A6"].font=ft2 
                sheet["A7"].font=ft2
                sheet["D6"].font=ft2
                cell.alignment = Alignment(horizontal='center', vertical='center')  
                sheet.append([])
                sheet.append(["","Right","Wrong","Not Attempt","MAX"])
                black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for label in ["E", "B", "C", "D","A"]:
                    idx = label + str(9)
                    sheet[idx].alignment = alignment1
                    sheet[idx].font = ft1
                    sheet[idx].border=black_border
                for i in range(3):
                    lis=[]
                    if(i==0):
                        lis=["No.","","","",""]
                    elif(i==1): 
                        lis=["Marking",right_per,wrong_per,0,""]
                    elif(i==2):
                        lis=["Total","","","","Absent"]
                    for label in ["A", "B", "C", "D","E"]:
                        sheet[label+str(i+10)]=lis[ord(label)-65]
                        sheet[label+str(i+10)].alignment = alignment1
                        sheet[label+str(i+10)].border=black_border 
                sheet["A10"].font=ft1
                sheet["A11"].font=ft1
                sheet["A12"].font=ft1
                sheet["D11"].font=ft1
                sheet["B11"].font=Font(name='Century', size=12,color="00008000")
                sheet["C11"].font=Font(name='Century', size=12,color="00FF0000")
                sheet["E12"].font=Font(name='Century', size=12,color="000000FF")
                wb.save(check)
    return d

def generate_concise(right_per,wrong_per):
    l = defaultdict(list)
    ans=[]
    dirpath = os.path.join(direc)
    if os.path.isdir(dirpath) == False:
        os.mkdir(direc)
    final_srt=os.path.join(direc,"marksheet")
    if os.path.isdir(final_srt)==False:
        os.mkdir(final_srt)
    with open(response_path,'r') as f:
        reader =csv.reader(f)
        flag=0
        for words in reader:
            if words[6]=="ANSWER":
                flag=1
                ans=words[7:]
                break

    # if no row with roll_num="ANSWER" exist
    if flag==0:            
        raise ValueError
    
    with open(master_path,'r') as f:
        reader=csv.reader(f)
        for words in reader:
            if words[0]!="roll":
              l[words[0]].append(False)
              l[words[0]].append(words[1])
    

    with open(response_path,'r') as f:
        reader =csv.reader(f)
        negative_list=[]
        status_ans=[]
        for words in reader: 
            right=0
            wrong=0
            not_attempt=0
            max=0
            roll_no=words[6]
            roll_no.upper()
            if roll_no=="Roll Number":
                continue
            l[roll_no][0]=True
            l[roll_no].append(words[1])
            l[roll_no].append(words[4])
            for i in range(len(ans)):
                    max=max+1
                    if words[i+7]=="":
                        not_attempt=not_attempt+1
                    elif ans[i]==words[i+7]:
                        right=right+1
                    else:
                        wrong=wrong+1
            status_ans.append("["+str(right)+", "+str(wrong)+", "+str(not_attempt)+"]")
            negative_list.append(str((right_per*right+wrong_per*wrong))+"/"+str(max*right_per))
    
    df=pd.read_csv(response_path)
    df.rename(columns={
        'Score': 'Google_Score'
    }, inplace=True)
    df.insert(loc=6, column='Score_After_Negative', value=negative_list)
    df.insert(loc=len(df.columns), column='statusAns', value=status_ans)
    
    for key, value in l.items():
        if value[0]==False:
         df = df.append({'Name': value[1],'Google_Score':"Absent",'Roll Number':key}, ignore_index=True) 
    destination = os.path.join(final_srt,"concise_marksheet.csv")
    df.to_csv(destination,index=False)
def send_email(d,right,wrong):
    print("Sending mails\n")
    server = smtplib.SMTP_SSL("smtp.gmail.com",465)
    server.login(params["gmail-user"],params["gmail-password"])
    final_srt=os.path.join(direc,"marksheet")
    right_part = "+"+str(right)+" for correct answer"
    wrong_part = str(wrong)+" for wrong answer"
    body = params['body']+right_part+', '+wrong_part
    for filename in os.listdir(final_srt):
        if filename!="concise_marksheet.csv":
            split_tup = os.path.splitext(filename)
            name_roll=split_tup[0]
            if d[name_roll][0] == True:
                print(d[name_roll][1])
                msg = MIMEMultipart()
                msg['To'] = ""
                msg['From'] = params["gmail-user"]
                msg['Subject'] = params['subject']
                msg.attach(MIMEText(body,'plain'))
                path = os.path.join(final_srt,filename)
                attachment = open(path,'rb')
                part = MIMEBase('application','octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',"attachment",filename=filename)
                msg.attach(part)
                if d[name_roll][2]:
                    msg.replace_header('To',d[name_roll][2])
                    text = msg.as_string()
                    try:
                        server.sendmail(params["gmail-user"],d[name_roll][2],text)
                    except:
                        continue
                if d[name_roll][3]:
                    msg.replace_header('To',d[name_roll][3])
                    text = msg.as_string()
                    try:
                        server.sendmail(params["gmail-user"],d[name_roll][3],text)
                    except:
                        continue
    server.quit()
@app.route("/", methods=["GET", "POST"])
def GUI():
    global master_flag,response_flag,ndir,pos,neg,ldir,flag1,btn1_flag,btn1,btn2,btn2_flag,btn3,notlast,cantsend,gen_pos,gen_neg,gen_list
    if request.method == "POST":
        notlast = 0
        btn1_flag = 0
        btn1 = 0
        btn2 = 0 
        btn3 = 0
        btn2_flag = 0
        cantsend = 0
        if request.files:
            file = request.files["csv-file"]
            if file.filename == "master_roll.csv":
                master_flag = 1
            if file.filename == "responses.csv":
                response_flag = 1
            flag1 = 0
            # print(file)
            file.save(os.path.join(app.config["input_file_loc"], file.filename))
            ldir = os.listdir(loc)
            ndir = len(ldir)
            if ndir<2:
                flag1 = 1
            return redirect(request.url)
        elif request.form:
            if request.form["btn"] == "btn1":
                btn1 = 1
                notlast = 1
            elif request.form["btn"] == "btn2":
                btn2 = 1
                notlast = 1
            elif request.form["btn"] == "btn3":
                btn3 = 1
            if btn3 == 1:
                path2 = "sample_output"
                dir2 = os.path.join(path2,"marksheet","ANSWER.xlsx")
                if os.path.isfile(dir2) == True:
                    try:
                        send_email(gen_list,gen_pos,gen_neg)
                    except:
                        cantsend = -1
                        
                else:
                    cantsend = 1
            if notlast == 1:
                pos = request.form["pos"]
                neg = request.form["neg"]
            return redirect(request.url)
    if flag1 == 0:
        if btn1 == 1:
            btn1_flag = 1
            try:
                gen_list = generate_marksheet(float(pos),float(neg))
            except:
                btn1_flag = -1
                # print(btn1_flag)
            #print(gen_list)
            gen_pos = pos
            gen_neg = neg
            
        elif btn2 == 1:
            btn2_flag = 1
            try:
                generate_concise(float(pos),float(neg))
            except:
                btn2_flag = -1
             
            
 
    return render_template('index.html',val1 = master_flag,val2 = response_flag,val3 = ndir,val4 = pos,val5 = neg,val6 = ldir, val7 = flag1, val8 = btn1_flag, val9 = btn2_flag, val10 = btn1, val11 = btn2, val12 = btn3, val13 = cantsend, val14 = btn3, val15 = gen_pos, val16 = gen_neg)

if __name__ == "__main__":
    app.run(debug=True)



